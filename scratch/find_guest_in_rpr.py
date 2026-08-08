import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

if 'Rpr' in wb.sheetnames:
    sheet = wb['Rpr']
    print(f"Rpr sheet max row: {sheet.max_row}")
    # Print headers (row 1 or 2)
    for r in range(1, 5):
        print(f"Row {r}: {[sheet.cell(row=r, column=col).value for col in range(1, 25)]}")
    
    # Search for guests in all rows
    guests = ["SELIN", "FATMA", "CAN", "OZGE", "FEHMI"]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, 20):
            val = sheet.cell(row=r, column=c).value
            if val:
                val_str = str(val).upper()
                for g in guests:
                    if g in val_str:
                        print(f"Found {g} in Rpr sheet at Row {r}, Col {c}: {val_str}")
else:
    print("Rpr sheet not found!")
