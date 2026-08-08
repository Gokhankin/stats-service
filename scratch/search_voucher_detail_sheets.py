import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# We want to find which sheet has voucher numbers or guest lists
vouchers_to_find = ["2573542", "2543948", "2584889", "2585155"]

for name in wb.sheetnames:
    sheet = wb[name]
    # let's look at a few rows or search
    found = False
    for r in range(1, min(100, sheet.max_row + 1)):
        for c in range(1, min(20, sheet.max_column + 1)):
            val = sheet.cell(row=r, column=c).value
            if val:
                val_str = str(val)
                for v in vouchers_to_find:
                    if v in val_str:
                        print(f"Voucher {v} found in Sheet: '{name}' at Row: {r}, Col: {c}")
                        found = True
                        break
        if found:
            break
