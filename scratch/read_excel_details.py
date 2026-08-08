import openpyxl
import pandas as pd

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Sheet Names in Excel:", wb.sheetnames)

# Let's inspect the first sheet
sheet = wb.active
print("Active Sheet Name:", sheet.title)

# Read first 15 rows of the active sheet to see its structure
data = []
for r in range(1, 20):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    data.append(row_vals)

for idx, d in enumerate(data):
    print(f"Row {idx+1}: {d}")
