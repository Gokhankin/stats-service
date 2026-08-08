import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

if 'Genel Forecast-R' in wb.sheetnames:
    sheet = wb['Genel Forecast-R']
    print(f"Genel Forecast-R sheet max row: {sheet.max_row}")
    for r in range(1, 20):
        print(f"Row {r}: {[sheet.cell(row=r, column=col).value for col in range(1, 15)]}")
else:
    print("Genel Forecast-R sheet not found!")
