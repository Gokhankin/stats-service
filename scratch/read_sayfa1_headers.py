import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheetname in ['Sayfa1', 'Sayfa2', 'Calc']:
    if sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"\n--- Sheet {sheetname} max row: {sheet.max_row} ---")
        for r in range(1, min(10, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=r, column=col).value for col in range(1, 15)]
            print(f"Row {r}: {row_vals}")
    else:
        print(f"Sheet {sheetname} not found!")
