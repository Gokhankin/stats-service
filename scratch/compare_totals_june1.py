import openpyxl
import sys
sys.path.append("/home/society/Masaüstü/Sedna Dashboard")
import queries as Q
import pyodbc
import pandas as pd

# 1. Read the new June 1st Excel Agency EUR revenues
excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -01.06.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Act Gelir Döküm']

excel_agencies = {}
# Read rows 4 to 100 (let's dynamically find the rows)
for r in range(4, 150):
    ac_name = sheet.cell(row=r, column=4).value
    # If we hit the Total row, we stop or skip
    if ac_name and str(ac_name).strip().lower() == 'total':
        break
    if ac_name:
        rooms = sheet.cell(row=r, column=11).value or 0
        pax = sheet.cell(row=r, column=12).value or 0
        try_rev = sheet.cell(row=r, column=17).value or 0
        eur_rev = sheet.cell(row=r, column=18).value or 0
        excel_agencies[ac_name.strip().upper()] = {
            'rooms': rooms,
            'pax': pax,
            'try': try_rev,
            'eur': eur_rev
        }

# 2. Get Live revenues from Sedna
conn_str = "DRIVER={ODBC Driver 18 for SQL Server};SERVER=192.168.0.41,1433;DATABASE=SednaAdakoy;UID=gokhan;PWD=Ad!!2025!!;TrustServerCertificate=yes;"
conn = pyodbc.connect(conn_str)
df_folio = Q.get_agency_folio_revenue(conn, 2026, include_april=False)

live_agencies = {}
for _, row in df_folio.iterrows():
    ac = str(row['acenta']).strip().upper()
    live_agencies[ac] = {
        'rooms': int(row['night_room']),
        'pax': int(row['night_pax']),
        'try': float(row['gelir_tl']),
        'eur': float(row['gelir_raw'])
    }

# Normalize names for comparison
excel_normalized = {}
name_mapping = {
    'WALK-IN EURO': 'MUNFERIT EURO',
    'WALK-IN GBP': 'MUNFERIT GBP',
    'WALK-IN TL': 'MUNFERIT TL',
    'WALK IN EURO': 'MUNFERIT EURO',
    'WALK IN GBP': 'MUNFERIT GBP',
    'WALK IN TL': 'MUNFERIT TL',
    'TUI DEUTSCHLAND': 'TUI DEUTCHLAND',
    'WEBBEDS': 'WEBBEDS-B2B',
    'DER TOUR-DEST': 'DER TOUR- DEST.',
    'HOTELBEDS-B2B': 'HOTELBEDS - B2B'
}

for k, v in excel_agencies.items():
    norm_k = name_mapping.get(k, k)
    if norm_k in excel_normalized:
        excel_normalized[norm_k]['rooms'] += v['rooms']
        excel_normalized[norm_k]['pax'] += v['pax']
        excel_normalized[norm_k]['try'] += v['try']
        excel_normalized[norm_k]['eur'] += v['eur']
    else:
        excel_normalized[norm_k] = v.copy()

print("--- Normalized Agency Differences (June 1st Excel vs Live SQL) ---")
print(f"{'Normalized Agency':<25} | {'Excel Rooms':<11} {'Live Rooms':<11} {'Room Diff':<9} | {'Excel EUR':<12} {'Live EUR':<12} {'EUR Diff':<10}")
print("-" * 110)

excel_total_eur = 0
live_total_eur = 0

all_keys = sorted(list(set(excel_normalized.keys()) | set(live_agencies.keys())))
for k in all_keys:
    ex = excel_normalized.get(k, {'rooms': 0, 'pax': 0, 'try': 0.0, 'eur': 0.0})
    lv = live_agencies.get(k, {'rooms': 0, 'pax': 0, 'try': 0.0, 'eur': 0.0})
    
    excel_total_eur += ex['eur']
    live_total_eur += lv['eur']
    
    room_diff = lv['rooms'] - ex['rooms']
    eur_diff = lv['eur'] - ex['eur']
    
    if room_diff != 0 or abs(eur_diff) > 0.01:
        print(f"{k:<25} | {ex['rooms']:<11} {lv['rooms']:<11} {room_diff:<9} | {ex['eur']:<12,.2f} {lv['eur']:<12,.2f} {eur_diff:<10,.2f}")

print("-" * 110)
print(f"{'TOTAL':<25} | {'-':<11} {'-':<11} {'-':<9} | {excel_total_eur:<12,.2f} {live_total_eur:<12,.2f} {live_total_eur - excel_total_eur:<10,.2f}")

conn.close()
wb.close()
