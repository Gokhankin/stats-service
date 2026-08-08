import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Act Gelir Döküm']

# Search for "TOPLAM" or "TOTAL"
for r in range(1, 150):
    for c in range(1, 10):
        val = sheet.cell(row=r, column=c).value
        if val and any(kw in str(val).upper() for kw in ["TOPLAM", "TOTAL"]):
            print(f"Row {r}: {[sheet.cell(row=r, column=col).value for col in range(1, 20)]}")
