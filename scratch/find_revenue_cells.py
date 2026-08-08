import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

for name in wb.sheetnames:
    sheet = wb[name]
    for r in range(1, 100):
        for c in range(1, 15):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val)
                if "1223758" in val_str or "64688696" in val_str or "64,688,696" in val_str:
                    print(f"Match found in Sheet: '{name}', Row: {r}, Col: {c} -> {val_str}")
