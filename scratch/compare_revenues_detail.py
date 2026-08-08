import openpyxl
import sys
sys.path.append("/home/society/Masaüstü/Sedna Dashboard")
import queries as Q
import pyodbc
import pandas as pd

# 1. Read Excel Agency EUR revenues
excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Act Gelir Döküm']

excel_agencies = {}
# Excel has columns: Col 4: AgencyName, Col 11: RoomNights, Col 12: BedNights, Col 17: TRY Revenue, Col 18: EUR Revenue
# We saw row 59 contains the total, so let's read rows 5 to 58.
for r in range(5, 59):
    ac_name = sheet.cell(row=r, column=4).value
    if ac_name:
        rooms = sheet.cell(row=r, column=11).value or 0
        pax = sheet.cell(row=r, column=12).value or 0
        try_rev = sheet.cell(row=r, column=17).value or 0
        eur_rev = sheet.cell(row=r, column=18).value or 0
        excel_agencies[ac_name.strip().upper()] = {
            'rooms': rooms,
            'pax': pax,
            'try': try_rev,
            'eur': eur_rev
        }

# 2. Get Live revenues from Sedna
conn_str = "DRIVER={ODBC Driver 18 for SQL Server};SERVER=192.168.0.41,1433;DATABASE=SednaAdakoy;UID=gokhan;PWD=Ad!!2025!!;TrustServerCertificate=yes;"
conn = pyodbc.connect(conn_str)
df_folio = Q.get_agency_folio_revenue(conn, 2026, include_april=False)

# Let's map agency codes / names from df_folio
# columns: acenta, night_room, night_pax, night_adult, gelir_raw, gelir_tl, commission, channel
live_agencies = {}
for _, row in df_folio.iterrows():
    ac = str(row['acenta']).strip().upper()
    live_agencies[ac] = {
        'rooms': int(row['night_room']),
        'pax': int(row['night_pax']),
        'try': float(row['gelir_tl']),
        'eur': float(row['gelir_raw'])
    }

# 3. Compare them!
print(f"{'AGENCY / SOURCE':<25} | {'EXCEL ROOMS':<11} {'LIVE ROOMS':<11} {'ROOM DIFF':<10} | {'EXCEL EUR':<12} {'LIVE EUR':<12} {'EUR DIFF':<10}")
print("-" * 105)

all_keys = sorted(list(set(excel_agencies.keys()) | set(live_agencies.keys())))
for k in all_keys:
    ex = excel_agencies.get(k, {'rooms': 0, 'pax': 0, 'try': 0.0, 'eur': 0.0})
    lv = live_agencies.get(k, {'rooms': 0, 'pax': 0, 'try': 0.0, 'eur': 0.0})
    
    room_diff = lv['rooms'] - ex['rooms']
    eur_diff = lv['eur'] - ex['eur']
    
    if room_diff != 0 or abs(eur_diff) > 0.01:
        print(f"{k:<25} | {ex['rooms']:<11} {lv['rooms']:<11} {room_diff:<10} | {ex['eur']:<12,.2f} {lv['eur']:<12,.2f} {eur_diff:<10,.2f}")

conn.close()
