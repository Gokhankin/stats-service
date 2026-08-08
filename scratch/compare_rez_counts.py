import openpyxl
import sys, pyodbc, pandas as pd
sys.path.append("/home/society/Masaüstü/Sedna Dashboard")
import queries as Q

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Act Gelir Döküm']

# Sütun yapısı:
# Col4: Acenta adı, Col5: Rez sayısı (sezon), Col11: Sezon oda geceleri, Col12: Sezon pax
excel_data = {}
for r in range(4, 59):
    ac_name = sheet.cell(row=r, column=4).value
    if ac_name and str(ac_name).strip() not in ('', 'Total '):
        rez_count   = sheet.cell(row=r, column=5).value or 0
        rooms_season = sheet.cell(row=r, column=11).value or 0
        pax_season   = sheet.cell(row=r, column=12).value or 0
        excel_data[ac_name.strip().upper()] = {
            'rez': int(rez_count),
            'rooms': int(rooms_season),
            'pax': int(pax_season)
        }

# Excel toplam satırı (Row 59)
excel_total_rez   = sheet.cell(row=59, column=5).value or 0
excel_total_rooms = sheet.cell(row=59, column=11).value or 0
excel_total_pax   = sheet.cell(row=59, column=12).value or 0

# Canlı DB
conn_str = "DRIVER={ODBC Driver 18 for SQL Server};SERVER=192.168.0.41,1433;DATABASE=SednaAdakoy;UID=gokhan;PWD=Ad!!2025!!;TrustServerCertificate=yes;"
conn = pyodbc.connect(conn_str)
df_folio = Q.get_agency_folio_revenue(conn, 2026, include_april=False)

live_data = {}
for _, row in df_folio.iterrows():
    ac = str(row['acenta']).strip().upper()
    live_data[ac] = {
        'rez':   int(row.get('rez', 0)),
        'rooms': int(row['night_room']),
        'pax':   int(row['night_pax'])
    }

# İsim normalizasyonu
name_map = {
    'WALK-IN EURO': 'MUNFERIT EURO',
    'WALK-IN GBP':  'MUNFERIT GBP',
    'WALK-IN TL':   'MUNFERIT TL',
    'WEBBEDS':      'WEBBEDS-B2B',
    'DER TOUR-DEST':'DER TOUR- DEST.',
    'HOTELBEDS-B2B':'HOTELBEDS - B2B',
}
excel_norm = {}
for k, v in excel_data.items():
    nk = name_map.get(k, k)
    if nk in excel_norm:
        excel_norm[nk]['rez']   += v['rez']
        excel_norm[nk]['rooms'] += v['rooms']
        excel_norm[nk]['pax']   += v['pax']
    else:
        excel_norm[nk] = v.copy()

print(f"\n{'ACENTA':<25} | {'EXCEL REZ':>9} {'LIVE REZ':>8} {'REZ FARK':>8} | {'EXCEL ODA':>9} {'LIVE ODA':>8} {'ODA FARK':>8}")
print("-" * 85)

ex_total_rez = 0
lv_total_rez = 0
ex_total_rooms = 0
lv_total_rooms = 0
fark_var = False

all_keys = sorted(set(excel_norm) | set(live_data))
for k in all_keys:
    ex = excel_norm.get(k, {'rez':0,'rooms':0,'pax':0})
    lv = live_data.get(k,  {'rez':0,'rooms':0,'pax':0})
    ex_total_rez   += ex['rez']
    lv_total_rez   += lv['rez']
    ex_total_rooms += ex['rooms']
    lv_total_rooms += lv['rooms']
    rdiff  = lv['rez']   - ex['rez']
    odiff  = lv['rooms'] - ex['rooms']
    if rdiff != 0 or odiff != 0:
        fark_var = True
        print(f"{k:<25} | {ex['rez']:>9} {lv['rez']:>8} {rdiff:>+8} | {ex['rooms']:>9} {lv['rooms']:>8} {odiff:>+8}")

print("-" * 85)
print(f"{'TOPLAM (hesaplanan)':<25} | {ex_total_rez:>9} {lv_total_rez:>8} {lv_total_rez - ex_total_rez:>+8} | {ex_total_rooms:>9} {lv_total_rooms:>8} {lv_total_rooms - ex_total_rooms:>+8}")
print(f"{'TOPLAM (Excel satırı)':<25} | {int(excel_total_rez):>9} {'':>8} {'':>8} | {int(excel_total_rooms):>9} {'':>8} {'':>8}")

if not fark_var:
    print("\n✅ Fark yok! Sadece TOURISTICA ve kur farklarından kaynaklanan farklar mevcut.")

conn.close()
