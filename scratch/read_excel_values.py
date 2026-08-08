import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Özet Rapor']

print("--- Print Özet Rapor Actual Values ---")
for r in range(1, 30):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    # filter out rows that are entirely None to keep output concise
    if any(x is not None for x in row_vals):
        print(f"Row {r}: {row_vals}")

print("\n--- Let us check the Sezon Tot from Excel ---")
# Let's search for "Sezon Tot" or similar in the sheet
for r in range(1, 100):
    for c in range(1, 10):
        val = sheet.cell(row=r, column=c).value
        if val and "Sezon" in str(val):
            print(f"Row {r}, Col {c}: {val} -> Row values: {[sheet.cell(row=r, column=col).value for col in range(1, 10)]}")
