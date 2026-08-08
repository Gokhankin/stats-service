import openpyxl

excel_path = "/home/society/Masaüstü/Sedna Dashboard/Club Adaköy Rezervasyon Detay Raporu -2026 Sezonu -25.05.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

if 'Act Gelir Döküm' in wb.sheetnames:
    sheet = wb['Act Gelir Döküm']
    print("--- Print 'Act Gelir Döküm' first 30 rows ---")
    for r in range(1, 40):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        if any(x is not None for x in row_vals):
            print(f"Row {r}: {row_vals}")
else:
    print("Act Gelir Döküm sheet NOT found!")
